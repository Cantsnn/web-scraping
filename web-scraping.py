import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font

current_file = load_workbook("products.xlsx") #Verilen dosyayi aciyoruz
current_excel = current_file.active
max_row = current_excel.max_row
link = "https://www.spx.com.tr"
urls=[]

for rows in range (1,max_row+1): #Verilen dosyadaki urlleri çekip listeye kaydediyoruz
    urls.append(link + current_excel["A{}".format(rows)].value)
    

new_file = Workbook() #Cektigimiz verileri kaydetmek icin yeni bir excel dosyasi olusturuyoruz
new_excel = new_file.active
new_excel.title = "Product Information"

#Sutun isimlerini döngü yardımıyla,kalin olacak sekilde yazdiriyoruz
table_titles=["Product Brand","Product Name","Product Price(TL)","Product Code","Product Availability","Product Urls"]
kalin= Font(bold=True)
for asciiNo in range(65,71):
    new_excel["{}".format(chr(asciiNo)+str(1))] = table_titles[asciiNo-65]
    new_excel["{}".format(chr(asciiNo)+str(1))].font = kalin
    
#Verilen sayfada veri tipine ve class sözlüğüne göre arama yapıp
#cıktılarını gene excel dosyasında verilen sutun ve satır bilgilerine kaydeden fonk
def getData (soup,dataType,classDict,colName,rowNo):
    
    inc_data = soup.find_all(dataType,classDict)
    
    product_info = inc_data[0].text
    product_info = product_info.replace("\n","")
    if(classDict["class"] == "product__price"): #urun fiyatlarini stringten sayiye cevirmek icin traslama islemi
        product_info = product_info.replace("TL","")   
        product_info = product_info.replace(",","")
        new_excel["{}".format(colName+str(rowNo+2))] = float(product_info)
    else:
        new_excel["{}".format(colName+str(rowNo+2))] = product_info
    
#Urunlerin stok sayilarini veren fonk.
def getStocks(soup,classDict):
    
    incoming_size = soup.find_all("a",classDict)   
    return len(incoming_size)


row = 0
while(row<max_row):

    page = requests.get(urls[row])#sayfalari cekiyor
    
    if(page.status_code!=200): #hata olusursa bir sonraki urle geciyor
        print("Sayfa Bulunamadi!!!")
        row+=1
        continue
    else:
        soup = BeautifulSoup(page.content,'html.parser') #sayfa icerigini parcaliyor
        new_excel["F{}".format(row+2)] = urls[row] #linkleri yeni dosyaya kaydediyor

        getData(soup,"a",{"class":"product__brand"},"A",row) #urun markalarini,
        getData(soup,"h1",{"class":"product__title"},"B",row) #isimlerini,      
        getData(soup,"div",{"class":"product__price"},"C",row) #fiyatlarini 
        getData(soup,"div",{"class":"product__code"},"D",row) #kodlarini cekiyoruz, tabloda hangi sutuna yazilacaginiveriyoruz.
        #bulunan model sayisini aliyoruz
        size_no = getStocks(soup,{"class":("d-flex align-items-center justify-content-center text-reset product__variant product__size-variant mb-3 js-variant","d-flex align-items-center justify-content-center text-reset product__variant product__size-variant mb-3 js-variant selected font-weight-bold")})
        if(size_no==0): #bulunan model sayisi yoksa "0" seklinde yazdırıyoruz
            new_excel["E{}".format(row+2)] = "% 0"
        else:#olmayan bedenleri alip yüzdeyi hesapliyoruz
            NoneSize_no = getStocks(soup,{"class":"d-flex align-items-center justify-content-center text-reset product__variant product__size-variant mb-3 js-variant disabled"})
            percent = size_no/(size_no+NoneSize_no)*100 
            new_excel["E{}".format(row+2)] = "% {}".format(round(percent,2))
    row+=1

#urunlerin toplamını gostermek icin tabloya komut ekliyor, bu kısmı tabloda enterlamak gerekiyor 
new_excel["C{}".format(row+2)] = "=TOPLA(C2:C{})".format(row+1)
new_file.save("output.xlsx") #olusturdugumuz dosyayi kaydediyoruz













